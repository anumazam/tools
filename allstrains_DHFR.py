import openpyxl

name = input("What is the name of your file? Enter with extension (.xlsx) and in single quotes. ")

wb = openpyxl.load_workbook(name, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')

# create new sheet for correctly formatted raw data
wb.create_sheet(index=2, title='raw data')
sheet2 = wb.get_sheet_by_name('raw data')

from openpyxl.utils import get_column_letter

for y in range(1, sheet.max_column + 1, 1):	
	sheet2.column_dimensions[get_column_letter(y)].width = 12
	y+=1

wb.save(name)

# populate raw data table
x=1
y=0

for j in range(2, sheet.max_column + 1, 1):
	for i in range(3, 18, 2):		
		y+=1
		sheet2.cell(row=y, column=x).value = sheet.cell(row=i, column=j).value

		# go to next column
		if y > 7:
			y=0
			x+=1

wb.save(name)

# un-jumble the plate in a new sheet
wb.create_sheet(index=3, title='sorted data & analysis')
sheet3 = wb.get_sheet_by_name('sorted data & analysis')

for y in range(1, 17, 1):	
	sheet3.column_dimensions[get_column_letter(y)].width = 14.5
	y+=1
	
sheet3.cell(row=1, column=1).value = "W103A +MEV"
sheet3.cell(row=2, column=1).value = sheet2['A1'].value
sheet3.cell(row=3, column=1).value = sheet2['C1'].value
sheet3.cell(row=4, column=1).value = sheet2['E1'].value
sheet3.cell(row=5, column=1).value = sheet2['G1'].value
sheet3.cell(row=6, column=1).value = sheet2['L1'].value
sheet3.cell(row=7, column=1).value = sheet2['L3'].value

sheet3.cell(row=1, column=2).value = "W103A -MEV"
sheet3.cell(row=2, column=2).value = sheet2['B5'].value
sheet3.cell(row=3, column=2).value = sheet2['D5'].value
sheet3.cell(row=4, column=2).value = sheet2['F5'].value
sheet3.cell(row=5, column=2).value = sheet2['H5'].value
sheet3.cell(row=6, column=2).value = sheet2['L5'].value
sheet3.cell(row=7, column=2).value = sheet2['L7'].value

sheet3.cell(row=1, column=3).value = "L78A +MEV"
sheet3.cell(row=2, column=3).value = sheet2['A2'].value
sheet3.cell(row=3, column=3).value = sheet2['C2'].value
sheet3.cell(row=4, column=3).value = sheet2['E2'].value
sheet3.cell(row=5, column=3).value = sheet2['G2'].value
sheet3.cell(row=6, column=3).value = sheet2['K1'].value
sheet3.cell(row=7, column=3).value = sheet2['K3'].value

sheet3.cell(row=1, column=4).value = "L78A -MEV"
sheet3.cell(row=2, column=4).value = sheet2['B6'].value
sheet3.cell(row=3, column=4).value = sheet2['D6'].value
sheet3.cell(row=4, column=4).value = sheet2['F6'].value
sheet3.cell(row=5, column=4).value = sheet2['H6'].value
sheet3.cell(row=6, column=4).value = sheet2['K5'].value
sheet3.cell(row=7, column=4).value = sheet2['K7'].value

sheet3.cell(row=1, column=5).value = "R134A +MEV"
sheet3.cell(row=2, column=5).value = sheet2['A3'].value
sheet3.cell(row=3, column=5).value = sheet2['C3'].value
sheet3.cell(row=4, column=5).value = sheet2['E3'].value
sheet3.cell(row=5, column=5).value = sheet2['G3'].value
sheet3.cell(row=6, column=5).value = sheet2['J1'].value
sheet3.cell(row=7, column=5).value = sheet2['J3'].value

sheet3.cell(row=1, column=6).value = "R134A -MEV"
sheet3.cell(row=2, column=6).value = sheet2['B7'].value
sheet3.cell(row=3, column=6).value = sheet2['D7'].value
sheet3.cell(row=4, column=6).value = sheet2['F7'].value
sheet3.cell(row=5, column=6).value = sheet2['H7'].value
sheet3.cell(row=6, column=6).value = sheet2['J5'].value
sheet3.cell(row=7, column=6).value = sheet2['J7'].value

sheet3.cell(row=1, column=7).value = "F286A +MEV"
sheet3.cell(row=2, column=7).value = sheet2['A4'].value
sheet3.cell(row=3, column=7).value = sheet2['C4'].value
sheet3.cell(row=4, column=7).value = sheet2['E4'].value
sheet3.cell(row=5, column=7).value = sheet2['G4'].value
sheet3.cell(row=6, column=7).value = sheet2['I1'].value
sheet3.cell(row=7, column=7).value = sheet2['I3'].value

sheet3.cell(row=1, column=8).value = "F286A -MEV"
sheet3.cell(row=2, column=8).value = sheet2['B8'].value
sheet3.cell(row=3, column=8).value = sheet2['D8'].value
sheet3.cell(row=4, column=8).value = sheet2['F8'].value
sheet3.cell(row=5, column=8).value = sheet2['H8'].value
sheet3.cell(row=6, column=8).value = sheet2['I5'].value
sheet3.cell(row=7, column=8).value = sheet2['I7'].value

sheet3.cell(row=1, column=9).value = "LIGSC +MEV"
sheet3.cell(row=2, column=9).value = sheet2['B1'].value
sheet3.cell(row=3, column=9).value = sheet2['D1'].value
sheet3.cell(row=4, column=9).value = sheet2['F1'].value
sheet3.cell(row=5, column=9).value = sheet2['H1'].value
sheet3.cell(row=6, column=9).value = sheet2['L2'].value
sheet3.cell(row=7, column=9).value = sheet2['L4'].value

sheet3.cell(row=1, column=10).value = "LIGSC -MEV"
sheet3.cell(row=2, column=10).value = sheet2['A5'].value
sheet3.cell(row=3, column=10).value = sheet2['C5'].value
sheet3.cell(row=4, column=10).value = sheet2['E5'].value
sheet3.cell(row=5, column=10).value = sheet2['G5'].value
sheet3.cell(row=6, column=10).value = sheet2['L6'].value
sheet3.cell(row=7, column=10).value = sheet2['L8'].value

sheet3.cell(row=1, column=11).value = "S9G10 +MEV"
sheet3.cell(row=2, column=11).value = sheet2['B2'].value
sheet3.cell(row=3, column=11).value = sheet2['D2'].value
sheet3.cell(row=4, column=11).value = sheet2['F2'].value
sheet3.cell(row=5, column=11).value = sheet2['H2'].value
sheet3.cell(row=6, column=11).value = sheet2['K2'].value
sheet3.cell(row=7, column=11).value = sheet2['K4'].value

sheet3.cell(row=1, column=12).value = "S9G10 -MEV"
sheet3.cell(row=2, column=12).value = sheet2['A6'].value
sheet3.cell(row=3, column=12).value = sheet2['C6'].value
sheet3.cell(row=4, column=12).value = sheet2['E6'].value
sheet3.cell(row=5, column=12).value = sheet2['G6'].value
sheet3.cell(row=6, column=12).value = sheet2['K6'].value
sheet3.cell(row=7, column=12).value = sheet2['K8'].value

sheet3.cell(row=1, column=13).value = "S9Y197A +MEV"
sheet3.cell(row=2, column=13).value = sheet2['B3'].value
sheet3.cell(row=3, column=13).value = sheet2['D3'].value
sheet3.cell(row=4, column=13).value = sheet2['F3'].value
sheet3.cell(row=5, column=13).value = sheet2['H3'].value
sheet3.cell(row=6, column=13).value = sheet2['J2'].value
sheet3.cell(row=7, column=13).value = sheet2['J4'].value

sheet3.cell(row=1, column=14).value = "S9Y197A -MEV"
sheet3.cell(row=2, column=14).value = sheet2['A7'].value
sheet3.cell(row=3, column=14).value = sheet2['C7'].value
sheet3.cell(row=4, column=14).value = sheet2['E7'].value
sheet3.cell(row=5, column=14).value = sheet2['G7'].value
sheet3.cell(row=6, column=14).value = sheet2['J6'].value
sheet3.cell(row=7, column=14).value = sheet2['J8'].value

sheet3.cell(row=1, column=15).value = "ISPA +MEV"
sheet3.cell(row=2, column=15).value = sheet2['B4'].value
sheet3.cell(row=3, column=15).value = sheet2['D4'].value
sheet3.cell(row=4, column=15).value = sheet2['F4'].value
sheet3.cell(row=5, column=15).value = sheet2['H4'].value
sheet3.cell(row=6, column=15).value = sheet2['I2'].value
sheet3.cell(row=7, column=15).value = sheet2['I4'].value

sheet3.cell(row=1, column=16).value = "ISPA -MEV"
sheet3.cell(row=2, column=16).value = sheet2['A8'].value
sheet3.cell(row=3, column=16).value = sheet2['C8'].value
sheet3.cell(row=4, column=16).value = sheet2['E8'].value
sheet3.cell(row=5, column=16).value = sheet2['G8'].value
sheet3.cell(row=6, column=16).value = sheet2['I6'].value
sheet3.cell(row=7, column=16).value = sheet2['I8'].value

wb.save(name)

# calculate averages and stdevs

sheet3.cell(row=10, column=1).value = '=AVERAGE(A1:A6)'
sheet3.cell(row=10, column=2).value = '=AVERAGE(B1:B6)'
sheet3.cell(row=10, column=3).value = '=AVERAGE(C1:C6)'
sheet3.cell(row=10, column=4).value = '=AVERAGE(D1:D6)'
sheet3.cell(row=10, column=5).value = '=AVERAGE(E1:E6)'
sheet3.cell(row=10, column=6).value = '=AVERAGE(F1:F6)'
sheet3.cell(row=10, column=7).value = '=AVERAGE(G1:G6)'
sheet3.cell(row=10, column=8).value = '=AVERAGE(H1:H6)'
sheet3.cell(row=10, column=9).value = '=AVERAGE(I1:I6)'
sheet3.cell(row=10, column=10).value = '=AVERAGE(J1:J6)'
sheet3.cell(row=10, column=11).value = '=AVERAGE(K1:K6)'
sheet3.cell(row=10, column=12).value = '=AVERAGE(L1:L6)'
sheet3.cell(row=10, column=13).value = '=AVERAGE(M1:M6)'
sheet3.cell(row=10, column=14).value = '=AVERAGE(N1:N6)'
sheet3.cell(row=10, column=15).value = '=AVERAGE(O1:O6)'
sheet3.cell(row=10, column=16).value = '=AVERAGE(P1:P6)'

sheet3.cell(row=11, column=1).value = '=STDEV(A1:A6)'
sheet3.cell(row=11, column=2).value = '=STDEV(B1:B6)'
sheet3.cell(row=11, column=3).value = '=STDEV(C1:C6)'
sheet3.cell(row=11, column=4).value = '=STDEV(D1:D6)'
sheet3.cell(row=11, column=5).value = '=STDEV(E1:E6)'
sheet3.cell(row=11, column=6).value = '=STDEV(F1:F6)'
sheet3.cell(row=11, column=7).value = '=STDEV(G1:G6)'
sheet3.cell(row=11, column=8).value = '=STDEV(H1:H6)'
sheet3.cell(row=11, column=9).value = '=STDEV(I1:I6)'
sheet3.cell(row=11, column=10).value = '=STDEV(J1:J6)'
sheet3.cell(row=11, column=11).value = '=STDEV(K1:K6)'
sheet3.cell(row=11, column=12).value = '=STDEV(L1:L6)'
sheet3.cell(row=11, column=13).value = '=STDEV(M1:M6)'
sheet3.cell(row=11, column=14).value = '=STDEV(N1:N6)'
sheet3.cell(row=11, column=15).value = '=STDEV(O1:O6)'
sheet3.cell(row=11, column=16).value = '=STDEV(P1:P6)'

wb.save(name)

# reformat for graphing
sheet3.cell(row=14, column=2).value = 'AVERAGE, +MEV'
sheet3.cell(row=14, column=3).value = 'AVERAGE, -MEV'
sheet3.cell(row=14, column=4).value = 'STDEV, +MEV'
sheet3.cell(row=14, column=5).value = 'STDEV, -MEV'

q=15
strains = ['W103A', 'L78A', 'R134A', 'F286A', 'LIGSC', 'S9G10', 'S9Y197A', 'ISPA']
for g in range(1, 9, 1):
	sheet3.cell(row=q, column=1).value = strains[g-1]
	q+=1

m=15
for k in range(1, sheet3.max_column + 1, 2):		
	sheet3.cell(row=m, column=2).value = sheet3.cell(row=10, column=k).value
	m+=1

m=15
for k in range(2, sheet3.max_column + 1, 2):		
	sheet3.cell(row=m, column=3).value = sheet3.cell(row=10, column=k).value
	m+=1
	
m=15
for k in range(1, sheet3.max_column + 1, 2):		
	sheet3.cell(row=m, column=4).value = sheet3.cell(row=11, column=k).value
	m+=1
	
m=15
for k in range(2, sheet3.max_column + 1, 2):		
	sheet3.cell(row=m, column=5).value = sheet3.cell(row=11, column=k).value
	m+=1

wb.save(name)

