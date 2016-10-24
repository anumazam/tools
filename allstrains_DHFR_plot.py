#!/usr/bin/python

import openpyxl
import numpy

name = input("What is the name of your file? Enter with extension (.xlsx) and in single quotes. ")

wb = openpyxl.load_workbook(name, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')

# create new sheet for correctly formatted raw data
wb.create_sheet(index=2, title='raw data')
sheet2 = wb.get_sheet_by_name('raw data')

from openpyxl.utils import get_column_letter

for y in range(1, sheet.max_column + 1, 1):	
	sheet2.column_dimensions[get_column_letter(y)].width = 12
	y+=1

wb.save(name)

# populate raw data table
x=1
y=0

for j in range(2, sheet.max_column + 1, 1):
	for i in range(3, 18, 2):		
		y+=1
		sheet2.cell(row=y, column=x).value = sheet.cell(row=i, column=j).value

		# go to next column
		if y > 7:
			y=0
			x+=1

wb.save(name)

# unjumble + plot data
	
l=[]
w=[]
lsd=[]
wsd=[]

a = sheet2['A1'].value
b = sheet2['C1'].value
c = sheet2['E1'].value
d = sheet2['G1'].value
e = sheet2['L1'].value
f = sheet2['L3'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

a = sheet2['B5'].value
b = sheet2['B5'].value
c = sheet2['F5'].value
d = sheet2['H5'].value
e = sheet2['L5'].value
f = sheet2['L7'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "L78A +MEV"
a = sheet2['A2'].value
b = sheet2['C2'].value
c = sheet2['E2'].value
d = sheet2['G2'].value
e = sheet2['K1'].value
f = sheet2['K3'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "L78A -MEV"
a = sheet2['B6'].value
b = sheet2['D6'].value
c = sheet2['F6'].value
d = sheet2['H6'].value
e = sheet2['K5'].value
f = sheet2['K7'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "R134A +MEV"
a = sheet2['A3'].value
b = sheet2['C3'].value
c = sheet2['E3'].value
d = sheet2['G3'].value
e = sheet2['J1'].value
f = sheet2['J3'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "R134A -MEV"
a = sheet2['B7'].value
b = sheet2['D7'].value
c = sheet2['F7'].value
d = sheet2['H7'].value
e = sheet2['J5'].value
f = sheet2['J7'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "F286A +MEV"
a = sheet2['A4'].value
b = sheet2['C4'].value
c = sheet2['E4'].value
d = sheet2['G4'].value
e = sheet2['I1'].value
f = sheet2['I3'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "F286A -MEV"
a = sheet2['B8'].value
b = sheet2['D8'].value
c = sheet2['F8'].value
d = sheet2['H8'].value
e = sheet2['I5'].value
f = sheet2['I7'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "LIGSC +MEV"
a = sheet2['B1'].value
b = sheet2['D1'].value
c = sheet2['F1'].value
d = sheet2['H1'].value
e = sheet2['L2'].value
f = sheet2['L4'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "LIGSC -MEV"
a = sheet2['A5'].value
b = sheet2['C5'].value
c = sheet2['E5'].value
d = sheet2['G5'].value
e = sheet2['L6'].value
f = sheet2['L8'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "S9G10 +MEV"
a = sheet2['B2'].value
b = sheet2['D2'].value
c = sheet2['F2'].value
d = sheet2['H2'].value
e = sheet2['K2'].value
f = sheet2['K4'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "S9G10 -MEV"
a = sheet2['A6'].value
b = sheet2['C6'].value
c = sheet2['E6'].value
d = sheet2['G6'].value
e = sheet2['K6'].value
f = sheet2['K8'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "S9Y197A +MEV"
a = sheet2['B3'].value
b = sheet2['D3'].value
c = sheet2['F3'].value
d = sheet2['H3'].value
e = sheet2['J2'].value
f = sheet2['J4'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "S9Y197A -MEV"
a = sheet2['A7'].value
b = sheet2['C7'].value
c = sheet2['E7'].value
d = sheet2['G7'].value
e = sheet2['J6'].value
f = sheet2['J8'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# "ISPA +MEV"
a = sheet2['B4'].value
b = sheet2['D4'].value
c = sheet2['F4'].value
d = sheet2['H4'].value
e = sheet2['I2'].value
f = sheet2['I4'].value
z= ((a+b+c+d+e+f)/6) 
l.append(z)
z= numpy.std([a,b,c,d,e,f])
lsd.append(z)

# "ISPA -MEV"
a = sheet2['A8'].value
b = sheet2['C8'].value
c = sheet2['E8'].value
d = sheet2['G8'].value
e = sheet2['I6'].value
f = sheet2['I8'].value
z= ((a+b+c+d+e+f)/6) 
w.append(z)
z= numpy.std([a,b,c,d,e,f])
wsd.append(z)

# plot.ly

from plotly import __version__
from plotly.offline import download_plotlyjs, init_notebook_mode, plot, iplot

from plotly.graph_objs import Bar, Scatter, Figure, Layout

# import plotly.plotly as py
import plotly.graph_objs as go

strains = ['W103A', 'L78A', 'R134A', 'F286A', 'LIGSC', 'S9G10', 'S9Y197A', 'ISPA']

trace1 = go.Bar(
    x=strains,
    y=l,    
    name='+mev',
	marker=dict(
		color='rgb(158,202,225)',
		line=dict(
			color='rgb(0,0,0)',
			width=2
			)
		),
    error_y=dict(
            type='data',
            array=lsd,
            visible=True
			),
		opacity=0.6
)

trace2 = go.Bar(
    x=strains,
    y=w,
    name='-mev',
    	marker=dict(
		color='rgb(201,051,102)',
		line=dict(
			color='rgb(0,0,0)',
			width=2
			)
		),
        error_y=dict(
            type='data',
            array=wsd,
            visible=True
			),
		opacity=0.6
)

data = [trace1, trace2]
layout = go.Layout(
	yaxis=dict(
		title='OD600',
		titlefont=dict(
			size=16
			),
		tickfont=dict(
			size=14
			)
		),
	xaxis=dict(
		titlefont=dict(
			size=16
			),
		tickfont=dict(
			size=14
			)
		),
    legend=dict(
        x=0,
        y=1.0,
    ),
    barmode='group'
)

fig = go.Figure(data=data, layout=layout)
plot(fig, filename='grouped-bar.html')

