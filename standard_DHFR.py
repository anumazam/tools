import openpyxl

name = input("What is the name of your file? Enter with extension (.xlsx) and in single quotes. ")

wb = openpyxl.load_workbook(name, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')

# create new sheet for correctly formatted raw data & analysis
wb.create_sheet(index=2, title='analysis')
sheet2 = wb.get_sheet_by_name('analysis')

from openpyxl.utils import get_column_letter

for y in range(1, sheet.max_column + 1, 1):	
	sheet2.column_dimensions[get_column_letter(y)].width = 14.5
	y+=1

wb.save(name)

# populate raw data table
x=1
y=0

for j in range(2, sheet.max_column + 1, 1):
	for i in range(3, 18, 2):		
		y+=1
		sheet2.cell(row=y, column=x).value = sheet.cell(row=i, column=j).value

		# go to next column
		if y > 7:
			y=0
			x+=1

wb.save(name)

# calculate averages and stdevs
sheet2.cell(row=10, column=1).value = '=AVERAGE(A1:A8)'
sheet2.cell(row=10, column=2).value = '=AVERAGE(B1:B8)'
sheet2.cell(row=10, column=3).value = '=AVERAGE(C1:C8)'
sheet2.cell(row=10, column=4).value = '=AVERAGE(D1:D8)'
sheet2.cell(row=10, column=5).value = '=AVERAGE(E1:E8)'
sheet2.cell(row=10, column=6).value = '=AVERAGE(F1:F8)'
sheet2.cell(row=10, column=7).value = '=AVERAGE(G1:G8)'
sheet2.cell(row=10, column=8).value = '=AVERAGE(H1:H8)'
sheet2.cell(row=10, column=9).value = '=AVERAGE(I1:I8)'
sheet2.cell(row=10, column=10).value = '=AVERAGE(J1:J8)'
sheet2.cell(row=10, column=11).value = '=AVERAGE(K1:K8)'
sheet2.cell(row=10, column=12).value = '=AVERAGE(L1:L8)'

sheet2.cell(row=11, column=1).value = '=STDEV(A1:A8)'
sheet2.cell(row=11, column=2).value = '=STDEV(B1:B8)'
sheet2.cell(row=11, column=3).value = '=STDEV(C1:C8)'
sheet2.cell(row=11, column=4).value = '=STDEV(D1:D8)'
sheet2.cell(row=11, column=5).value = '=STDEV(E1:E8)'
sheet2.cell(row=11, column=6).value = '=STDEV(F1:F8)'
sheet2.cell(row=11, column=7).value = '=STDEV(G1:G8)'
sheet2.cell(row=11, column=8).value = '=STDEV(H1:H8)'
sheet2.cell(row=11, column=9).value = '=STDEV(I1:I8)'
sheet2.cell(row=11, column=10).value = '=STDEV(J1:J8)'
sheet2.cell(row=11, column=11).value = '=STDEV(K1:K8)'
sheet2.cell(row=11, column=12).value = '=STDEV(L1:L8)'

wb.save(name)

# reformat for graphing
sheet2.cell(row=14, column=1).value = 'AVERAGE, +MEV'
sheet2.cell(row=14, column=2).value = 'AVERAGE, -MEV'
sheet2.cell(row=14, column=3).value = 'STDEV, +MEV'
sheet2.cell(row=14, column=4).value = 'STDEV, -MEV'

m=15
for k in range(1, sheet.max_column + 1, 2):		
	sheet2.cell(row=m, column=1).value = sheet2.cell(row=10, column=k).value
	m+=1

m=15
for k in range(2, sheet.max_column + 1, 2):		
	sheet2.cell(row=m, column=2).value = sheet2.cell(row=10, column=k).value
	m+=1
	
m=15
for k in range(1, sheet.max_column + 1, 2):		
	sheet2.cell(row=m, column=3).value = sheet2.cell(row=11, column=k).value
	m+=1
	
m=15
for k in range(2, sheet.max_column + 1, 2):		
	sheet2.cell(row=m, column=4).value = sheet2.cell(row=11, column=k).value
	m+=1

wb.save(name)

