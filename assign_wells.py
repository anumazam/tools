#!/usr/bin/python
import random

# this is a tool for randomly assigning wells for 8 strains on a 96 well plate.

# create vectors for rows and columns
columns = []
for n in range(1, 13, 1):
    columns.append(n)

rowIDs = [ 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H' ]

# top half of the plate (+mev)
board = [1]*6 + [2]*6 + [3]*6 + [4]*6 +[5]*6 + [6]*6 +[7]*6 + [8]*6
random.shuffle(board)
board = [board[i:i + 12] for i in xrange(0, 48, 12)]

A = board[0]
B = board[1]
C = board[2]
D = board[3]

# bottom half of the plate (-mev)
board = [1]*6 + [2]*6 + [3]*6 + [4]*6 +[5]*6 + [6]*6 +[7]*6 + [8]*6
random.shuffle(board)
board = [board[i:i + 12] for i in xrange(0, 48, 12)]

E = board[0]
F = board[1]
G = board[2]
H = board[3]

# create table for plate
print "\n"
from tabulate import tabulate
print tabulate ([ A, B, C, D, E, F, G, H ], headers = columns, showindex = rowIDs, numalign = "center")

allrows = [ A, B, C, D, E, F, G, H ]

# print all locations for each strain

for strain in range (0, 8, 1):
    print "\n wells for strain #" + str(strain+1) + ":"
    
    for n in range(0, 8, 1):
        for i, j in enumerate(allrows[n]):
            if j == strain+1:
                print rowIDs[n] + str(i+1),

print "\n"

# export the table to an excel spreadsheet
import openpyxl
import datetime
import csv

today = datetime.date.today()
filename = str(today) + ".csv"

fl = open(filename, 'w')

writer = csv.writer(fl)
for values in allrows:
    writer.writerow(values)

fl.close()

# clean up raw data
name = input("What is the name of your file? Enter with extension (.xlsx) and in single quotes. ")

wb = openpyxl.load_workbook(name, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')

# create new sheet for correctly formatted raw data
wb.create_sheet(index=2, title='raw data')
sheet2 = wb.get_sheet_by_name('raw data')

from openpyxl.utils import get_column_letter

for y in range(1, sheet.max_column + 1, 1):
    sheet2.column_dimensions[get_column_letter(y)].width = 12
    y+=1

wb.save(name)

# populate raw data table
x=1
y=0

for j in range(2, sheet.max_column + 1, 1):
    for i in range(3, 18, 2):
        y+=1
        sheet2.cell(row=y, column=x).value = sheet.cell(row=i, column=j).value
                
        # go to next column
        if y > 7:
            y=0
            x+=1

wb.save(name)

# assign data to new vectors for sorting

MEVstrains = ['W103A_mev', 'L78A_mev', 'R134A_mev', 'F286A_mev', 'LIGSC_mev', 'S9G10_mev', 'S9Y197A_mev', 'ISPA_mev']
noMEVstrains = ['W103A_nomev', 'L78A_nomev', 'R134A_nomev', 'F286A_nomev', 'LIGSC_nomev', 'S9G10_nomev', 'S9Y197A_nomev', 'ISPA_nomev']

#tuple(sheet2['A1': 'L8'])
#for rowOfCellObjects in sheet2['A1' : 'L8']:
#    for cellObj in rowOfCellObjects:
#        print (cellObj.coordinate, cellObj.value)
#    print ('--- END OF ROW ---')

W103A_mev = []
L78A_mev = []
R134A_mev = []
F286A_mev = []
LIGSC_mev = []
S9G10_mev = []
S9Y197A_mev = []
ISPA_mev = []
W103A_nomev = []
L78A_nomev = []
R134A_nomev = []
F286A_nomev = []
LIGSC_nomev = []
S9G10_nomev = []
S9Y197A_nomev = []
ISPA_nomev = []

rawdata = []
tuple(sheet2['A1': 'L8'])


for rowOfCellObjects in sheet2['A1' : 'A12']:
    for cellObj in rowOfCellObjects:
        for p in range(0, 11, 1):
            if allrows[0][p] == 1:
                W103A_mev.append(cellObj.value)
                print cellObj.value
            else:
                rawdata.append(cellObj.value)

print W103A_mev
print "\n"

print rawdata
print "\n"

print len(rawdata)